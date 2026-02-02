import os
import re
import logging
from datetime import datetime
import argparse
import pandas as pd
from dotenv import load_dotenv

from playwright.sync_api import (
    sync_playwright,
    TimeoutError as PWTimeoutError,
    Error as PWError,
)

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ======================
# НАСТРОЙКИ
# ======================
load_dotenv()

BASE_URL = os.getenv("BITRIX_BASE_URL", "https://globaldrive.ru").rstrip("/")
ENTITY_ID = os.getenv("ENTITY_ID", "4")

# --- Путь к реальному файлу с ID (вне репозитория) ---
EXTERNAL_IDS_PATH = r"C:\work_data\bitrix_ids\ids.csv"

# Если внешний файл есть — берём его,
# если нет — используем локальный пример
if os.path.exists(EXTERNAL_IDS_PATH):
    INPUT_FILE = EXTERNAL_IDS_PATH
else:
    INPUT_FILE = "ids_example.csv"
    print("⚠ Используется ids_example.csv — реальный ids.csv не найден")
OUTPUT_FILE = "bitrix_2025_report.xlsx"
LOG_FILE = "run.log"
SCREEN_DIR = "screenshots"

OK_YEAR = 2025

YANDEX_EXE = r"C:\Program Files\Yandex\YandexBrowser\Application\browser.exe"
YANDEX_USER_DATA = r"C:\Users\user\AppData\Local\Yandex\YandexBrowser\User Data - Playwright"
YANDEX_PROFILE_DIR = "Default"  # если будет другой профиль — поменяем на "Profile 1" и т.п.

# ======================
# ЛОГИРОВАНИЕ
# ======================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[logging.FileHandler(LOG_FILE, encoding="utf-8"), logging.StreamHandler()],
)

DATE_RE = re.compile(r"\d{2}\.\d{2}\.\d{4}\s+\d{2}:\d{2}:\d{2}")


def ensure_dirs():
    os.makedirs(SCREEN_DIR, exist_ok=True)


def make_url(item_id: str) -> str:
    # Эквивалент “ввела ID + нажала Найти”
    return (
        f"{BASE_URL}/bitrix/admin/highloadblock_rows_list.php"
        f"?PAGEN_1=1&SIZEN_1=20&ENTITY_ID={ENTITY_ID}&lang=ru"
        f"&set_filter=Y&adm_filter_applied=0&find_id={item_id}"
    )


def save_screenshot(page, status: str, item_id: str) -> str:
    safe_status = status.replace(" ", "_")
    path = os.path.join(SCREEN_DIR, f"{safe_status}_{item_id}.png")
    try:
        page.screenshot(path=path, full_page=True)
    except Exception:
        # если вкладка уже закрылась — скрин не получится
        return ""
    return path


def extract_year(text: str) -> tuple[str, int | None]:
    """Возвращает (дата_строкой, год) из текста строки таблицы."""
    text = (text or "").strip()
    m = DATE_RE.search(text)
    if not m:
        return "", None

    date_str = m.group(0)
    try:
        year = datetime.strptime(date_str, "%d.%m.%Y %H:%M:%S").year
        return date_str, year
    except Exception:
        return date_str, None


def is_logged_in_admin(page) -> bool:
    # В админке Bitrix обычно есть ссылка выхода вида ?logout=Y
    return page.locator("a[href*='logout=Y']").count() > 0

def is_login_page(page) -> bool:
    # Страница логина: есть оба поля ввода
    return (
        page.locator("input[name='USER_LOGIN']").count() > 0
        and page.locator("input[name='USER_PASSWORD']").count() > 0
    )

def ensure_admin_session(page):
    """
    Если мы НЕ в админке и видим форму логина — просим войти.
    Если мы в админке — ничего не делаем.
    """
    if is_logged_in_admin(page):
        return

    if is_login_page(page):
        print("⚠️ Открылась страница логина. Похоже, сессия закончилась.")
        input("👉 Войди в админку в этом окне и нажми ENTER, чтобы продолжить...")
        return

    # Если ни логина, ни админки — просто подождём чуть-чуть (страница могла не догрузиться)
    page.wait_for_timeout(500)



def wait_for_table(page):
    """Ждём появления таблицы в списке HL-блока."""
    try:
        page.wait_for_selector("table.adm-list-table", timeout=15000)
    except PWTimeoutError:
        # возможно нас выкинуло на логин
        ensure_admin_session(page)
        page.wait_for_selector("table.adm-list-table", timeout=15000)


def colorize_excel(path: str):
    """Красим строки по статусу + делаем шапку жирной + автоширина колонок."""
    wb = load_workbook(path)
    ws = wb.active

    fill_ok = PatternFill("solid", fgColor="C6EFCE")      # зелёный
    fill_fail = PatternFill("solid", fgColor="FFC7CE")    # красный
    fill_nf = PatternFill("solid", fgColor="FFEB9C")      # жёлтый
    fill_err = PatternFill("solid", fgColor="D9D9D9")     # серый

    for cell in ws[1]:
        cell.font = Font(bold=True)

    headers = [c.value for c in ws[1]]
    try:
        status_col = headers.index("Статус") + 1
    except ValueError:
        wb.save(path)
        return

    for r in range(2, ws.max_row + 1):
        status = ws.cell(row=r, column=status_col).value
        if status == "OK":
            fill = fill_ok
        elif status == "FAIL":
            fill = fill_fail
        elif status == "NOT FOUND":
            fill = fill_nf
        else:
            fill = fill_err

        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill

    for c in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(c)
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 70)

    wb.save(path)
EXTERNAL_IDS_PATH = r"C:\work_data\bitrix_ids\ids.csv"
EXAMPLE_IDS_FILE = "ids_example.csv"


def parse_args():
    parser = argparse.ArgumentParser(
        description="Bitrix HighloadBlock ID Checker (year validation)"
    )
    parser.add_argument(
        "--prod",
        action="store_true",
        help="Использовать внешний файл с коммерческими ID (EXTERNAL_IDS_PATH)",
    )
    parser.add_argument(
        "--example",
        action="store_true",
        help="Использовать пример ids_example.csv (для демо/портфолио)",
    )
    parser.add_argument(
        "--start-from",
        type=int,
        default=1,
        help="Начать проверку с N-й строки (нумерация с 1). По умолчанию 1.",
    )
    return parser.parse_args()


def resolve_input_file(args) -> str:
    """
    Выбирает, откуда брать IDs:
    - --example -> ids_example.csv
    - --prod -> EXTERNAL_IDS_PATH
    - по умолчанию: если EXTERNAL_IDS_PATH существует -> он, иначе example
    """
    if args.example and args.prod:
        raise ValueError("Нельзя одновременно использовать --example и --prod")

    if args.example:
        return EXAMPLE_IDS_FILE

    if args.prod:
        return EXTERNAL_IDS_PATH

    # auto
    return EXTERNAL_IDS_PATH if os.path.exists(EXTERNAL_IDS_PATH) else EXAMPLE_IDS_FILE


def load_ids_from_csv(path: str) -> list[str]:
    """
    Загружает IDs из CSV:
    - если есть колонка ID -> берёт её
    - иначе берёт первый столбец
    - чистит пустые/NaN/пробелы
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"Файл не найден: {path}")

    df = pd.read_csv(path)

    if df.empty:
        raise ValueError(f"Файл пустой: {path}")

    if "ID" in df.columns:
        raw = df["ID"]
    else:
        raw = df.iloc[:, 0]

    ids = raw.astype(str).str.strip().tolist()
    ids = [x for x in ids if x and x.lower() != "nan"]

    if len(ids) == 0:
        raise ValueError(f"В файле нет валидных ID: {path}")

    return ids


def main():
    ensure_dirs()

    args = parse_args()
input_file = resolve_input_file(args)

ids = load_ids_from_csv(input_file)

# старт с нужной строки
start_from = max(args.start_from, 1)
ids = ids[start_from - 1 :]

logging.info(f"IDs source: {input_file}")
logging.info(f"IDs loaded: {len(ids)} (start from {start_from})")

# если запущен демо-режим, напомним, что это пример
if input_file == EXAMPLE_IDS_FILE:
    print("ℹ Запущен DEMO режим (ids_example.csv). Для рабочих проверок используй --prod.")




    results = []

    with sync_playwright() as p:
        # Запуск Яндекс.Браузера с копией профиля
        context = p.chromium.launch_persistent_context(
            user_data_dir=YANDEX_USER_DATA,
            executable_path=YANDEX_EXE,
            headless=False,
            args=[f"--profile-directory={YANDEX_PROFILE_DIR}"],
        )
        page = context.new_page()

        # 2) Открываем админку
        page.goto(f"{BASE_URL}/bitrix/admin/", wait_until="domcontentloaded")
        input("👉 Если админка открылась и ты залогинена — нажми ENTER (если нет — войди и нажми ENTER)...")

        # 3) Проверяем IDs
        for i, item_id in enumerate(ids, start=1):
            item_id = str(item_id).strip()
            url = make_url(item_id)

            status = "FAIL"
            date_text = ""
            year = None
            screenshot_path = ""
            comment = ""

            try:
                page.goto(url, wait_until="domcontentloaded")

                # # сначала дождались таблицы
                wait_for_table(page)

                # потом проверили, что мы не на логине
                ensure_admin_session(page)

                # ИЩЕМ СТРОКУ ПО КОНКРЕТНОМУ ID 
                id_link = page.locator(f"a:has-text('{item_id}')").first

                if id_link.count() == 0:
                    status = "NOT FOUND"
                    comment = "ID не найден в таблице (фильтр/результат пустой)"
                    screenshot_path = save_screenshot(page, status, item_id)
                else:
                    row = id_link.locator("xpath=ancestor::tr[1]")
                    row_text = row.inner_text()

                    date_text, year = extract_year(row_text)

                    if year == OK_YEAR:
                        status = "OK"
                    else:
                        status = "FAIL"
                        comment = f"Ожидали {OK_YEAR}, фактически: {year if year else 'не распознано'}"
                        screenshot_path = save_screenshot(page, status, item_id)

            except PWTimeoutError as e:
                status = "ERROR"
                comment = f"Timeout: {e}"
                screenshot_path = save_screenshot(page, status, item_id)

            except PWError as e:
                # сюда попадают TargetClosedError и прочие ошибки Playwright
                status = "ERROR"
                comment = f"Playwright error (возможен вылет вкладки/сессии): {e}"
                screenshot_path = save_screenshot(page, status, item_id)

                # если вкладка/сессия упала — попробуем открыть админку заново
                try:
                    page.goto(f"{BASE_URL}/bitrix/admin/", wait_until="domcontentloaded")
                    ensure_admin_session(page)
                except Exception:
                    pass

            except Exception as e:
                status = "ERROR"
                comment = f"Exception: {e}"
                screenshot_path = save_screenshot(page, status, item_id)

            results.append(
                {
                    "ID": item_id,
                    "URL": url,
                    "Дата добавления": date_text,
                    "Год": year,
                    "Ожидаемый год": OK_YEAR,
                    "Статус": status,
                    "Комментарий": comment,
                    "Screenshot": screenshot_path,
                }
            )

            logging.info(f"[{i}/{len(ids)}] ID={item_id} -> {status} | year={year} | {comment}")

        context.close()

    # 4) Excel + раскраска
    pd.DataFrame(results).to_excel(OUTPUT_FILE, index=False)
    colorize_excel(OUTPUT_FILE)

    logging.info(f"✅ Done. Report: {OUTPUT_FILE}")
    logging.info(f"📝 Log: {LOG_FILE}")
    logging.info(f"📷 Screens: {SCREEN_DIR}/")


if __name__ == "__main__":
    main()
